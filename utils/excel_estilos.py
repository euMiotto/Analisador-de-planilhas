from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def obter_letra_coluna(worksheet, nome_col_header):
    
    # Cabeçalho está na primeira linha da planilha salva
    for col_idx in range(1, worksheet.max_column + 1):
        celula_cabecalho = worksheet.cell(row=1, column=col_idx).value
        if celula_cabecalho == nome_col_header:
            return worksheet.cell(row=1, column=col_idx).column_letter
    return None

def aplicar_destaques_excel(caminho_arquivo_excel, df_com_flags, colunas_para_destacar):
    
    try:
        wb = load_workbook(caminho_arquivo_excel)
        ws = wb.active
    except Exception as e:
        print(f"❌ ERRO ao reabrir '{caminho_arquivo_excel}' com openpyxl para estilização: {e}")
        return False

    fill_vermelho = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    pintou_algo = False

    # Mapeamento de letras
    letras_colunas_dados = {}
    for nome_col_dado in colunas_para_destacar.keys():
        letra = obter_letra_coluna(ws, nome_col_dado)
        if letra:
            letras_colunas_dados[nome_col_dado] = letra
        else:
            print(f"⚠️ ATENÇÃO: Não foi possível encontrar a letra da coluna '{nome_col_dado}' para aplicar estilos.")

    # A planilha salva com pandas
    mapeamento_indice_para_linha_excel = {
        original_idx: nova_linha_excel
        for nova_linha_excel, original_idx in enumerate(df_com_flags.index, start=2)
    }

    for idx_original_df, linha_dados_com_flags in df_com_flags.iterrows():
        linha_excel_correspondente = mapeamento_indice_para_linha_excel.get(idx_original_df)

        if linha_excel_correspondente is None:
            print(f"⚠️ AVISO: Não foi possível encontrar a linha do Excel para o índice {idx_original_df} do DataFrame durante a estilização.")
            continue

        for nome_col_dado, nome_col_flag in colunas_para_destacar.items():
            letra_col_dado = letras_colunas_dados.get(nome_col_dado)
            if letra_col_dado and nome_col_flag in linha_dados_com_flags and linha_dados_com_flags[nome_col_flag]:
                try:
                    ws[f'{letra_col_dado}{linha_excel_correspondente}'].fill = fill_vermelho
                    pintou_algo = True
                except Exception as e_cell:
                    print(f"❌ ERRO ao tentar pintar célula {letra_col_dado}{linha_excel_correspondente}: {e_cell}")


    if not pintou_algo and any(df_com_flags[flag].any() for flag in colunas_para_destacar.values() if flag in df_com_flags):
        print("⚠️ ATENÇÃO: Havia dados inválidos, mas parece que nada foi pintado. Verifique as letras das colunas e a lógica de pintura.")
    elif not pintou_algo:
        print("ℹ️ INFO: Nenhum dado inválido encontrado ou nenhuma célula foi pintada durante a estilização.")

    try:
        wb.save(caminho_arquivo_excel)
        print(f"✅ Estilos aplicados e planilha salva em: {caminho_arquivo_excel}")
        return True
    except Exception as e:
        print(f"❌ ERRO ao salvar o arquivo Excel final com estilos: {e}")
        return False
